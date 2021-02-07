from openpyxl import load_workbook

excelfile = load_workbook('fruit.xlsx')
sheet = excelfile.active

print(sheet['B1'].value)

alltitle = []
allprice = []
alldata = []

for i in range(3,7):
	pdname = sheet.cell(row=i,column=2).value
	pdprice = sheet.cell(row=i,column=3).value
	pddetail = sheet.cell(row=i,column=4).value
	alltitle.append(pdname)
	allprice.append(pdprice)
	alldata.append(pddetail)



from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time

driverpath = r'C:\Users\Uncle Engineer\Desktop\Python Bootcamp 2021\Class Main\EP3\chromedriver_win32\chromedriver.exe'
driver = webdriver.Chrome(driverpath)

url = 'http://www.uncle-machine.com/login/'

driver.get(url)


username = driver.find_element_by_id('username')
username.send_keys('loong9999@gmail.com')

password = driver.find_element_by_id('password')
password.send_keys('1234')
password.send_keys(Keys.RETURN) # press enter after type password

#button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
#button.click()

addurl = 'http://www.uncle-machine.com/addproduct/'

driver.get(addurl) # goto addurl



for pn,pp,pd in zip(alltitle,allprice,alldata):

	pdname = driver.find_element_by_id('name')
	pdprice = driver.find_element_by_id('price')
	pddetail = driver.find_element_by_id('detail')
	

	pd1_name = pn
	pd1_price = pp
	pd1_detail = pd
	


	pdname.send_keys(pd1_name)
	pdprice.send_keys(pd1_price)
	pddetail.send_keys(pd1_detail)
	
	time.sleep(5)

	button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
	button.click()
